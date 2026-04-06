# estimate_excel.py — генерація Excel кошторису (3 листи)

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Кольори ──
CLR_HEADER_BLUE  = "1F4E79"   # темно-синій — заголовки секцій
CLR_HEADER_LIGHT = "D6E4F0"   # світло-синій — шапка таблиці
CLR_TOTAL_ROW    = "FFF2CC"   # жовтий — рядки підсумків
CLR_GRAND_TOTAL  = "E2EFDA"   # зелений — фінальний підсумок
CLR_WHITE        = "FFFFFF"
CLR_LIGHT_GRAY   = "F5F5F5"


def _border(style="thin"):
    s = Side(style=style, color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(color):
    return PatternFill("solid", fgColor=color)


def _font(bold=False, size=11, color="000000", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _num_fmt(ws, cell, fmt):
    ws[cell].number_format = fmt


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _section_header(ws, row, col_start, col_end, text, color=CLR_HEADER_BLUE):
    """Злитий рядок-заголовок секції"""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(bold=True, size=11, color=CLR_WHITE, name="Arial")
    cell.fill = _header_fill(color)
    cell.alignment = _align("left")
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )


def _table_header(ws, row, cols: list):
    """Рядок шапки таблиці"""
    for i, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=text)
        cell.font = _font(bold=True, size=10, color="1F4E79")
        cell.fill = _header_fill(CLR_HEADER_LIGHT)
        cell.alignment = _align("center", wrap=True)
        cell.border = _border()


def _data_row(ws, row, values, is_alt=False):
    """Звичайний рядок даних"""
    fill = _header_fill(CLR_LIGHT_GRAY) if is_alt else _header_fill(CLR_WHITE)
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = _font(size=10)
        cell.fill = fill
        cell.border = _border()
        if isinstance(val, float) or isinstance(val, int):
            cell.alignment = _align("right")
        else:
            cell.alignment = _align("left", wrap=True)


def _total_row(ws, row, values, color=CLR_TOTAL_ROW):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = _font(bold=True, size=10)
        cell.fill = _header_fill(color)
        cell.border = _border()
        if isinstance(val, (float, int)) and not isinstance(val, bool):
            cell.alignment = _align("right")
        else:
            cell.alignment = _align("left")


# ════════════════════════════════════════════════════════════
# ЛИСТ 1 — МАТЕРІАЛИ
# ════════════════════════════════════════════════════════════
def _sheet_materials(wb, result):
    ws = wb.create_sheet("Матеріали")
    rate = result["usd_rate"]

    # Заголовок документу
    ws.merge_cells("A1:F1")
    ws["A1"] = f"КОШТОРИС — МАТЕРІАЛИ | {result['object_name']}"
    ws["A1"].font = Font(bold=True, size=13, color=CLR_HEADER_BLUE, name="Arial")
    ws["A1"].alignment = _align("center")

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Система: {result['system_name']} {result['insulation_thick']}мм | "
        f"Фарба: {result['paint_name']} | "
        f"Площа фасаду: {result['facade_area']} м² | "
        f"Курс: 1 USD = {rate:.2f} грн"
    )
    ws["A2"].font = _font(size=9, color="555555")
    ws["A2"].alignment = _align("center")

    # Шапка таблиці (рядок 4)
    _table_header(ws, 4, [
        "№", "Найменування матеріалу", "К-сть", "Од.", "Ціна USD", "Сума USD", "Сума грн"
    ])

    row = 5
    total_uah = 0.0
    total_usd = 0.0

    for i, mat in enumerate(result["materials"], 1):
        is_alt = (i % 2 == 0)
        _data_row(ws, row, [
            i,
            mat["name"],
            mat["qty"],
            mat["unit"],
            round(mat["price_usd"], 2),
            round(mat["total_usd"], 2),
            round(mat["total_uah"], 2),
        ], is_alt)
        # Формат чисел
        for col in [3, 5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        total_usd += mat["total_usd"]
        total_uah += mat["total_uah"]
        row += 1

    # Підсумок
    _total_row(ws, row, [
        "", "РАЗОМ МАТЕРІАЛИ", "", "", "",
        round(total_usd, 2), round(total_uah, 2)
    ], CLR_TOTAL_ROW)
    for col in [6, 7]:
        ws.cell(row=row, column=col).number_format = '#,##0.00'

    # Ширини колонок
    _set_col_widths(ws, {"A": 5, "B": 45, "C": 10, "D": 14, "E": 12, "F": 13, "G": 14})
    ws.row_dimensions[4].height = 30

    return ws


# ════════════════════════════════════════════════════════════
# ЛИСТ 2 — РОБОТИ
# ════════════════════════════════════════════════════════════
def _sheet_works(wb, result):
    ws = wb.create_sheet("Роботи")
    rate = result["usd_rate"]

    ws.merge_cells("A1:F1")
    ws["A1"] = f"КОШТОРИС — РОБОТИ | {result['object_name']}"
    ws["A1"].font = Font(bold=True, size=13, color=CLR_HEADER_BLUE, name="Arial")
    ws["A1"].alignment = _align("center")

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Площа фасаду: {result['facade_area']} м² | "
        f"Відкоси: {result['reveal_area']} м² | "
        f"Ліси: {'так' if result['need_scaffold'] else 'ні'} | "
        f"Курс: 1 USD = {rate:.2f} грн"
    )
    ws["A2"].font = _font(size=9, color="555555")
    ws["A2"].alignment = _align("center")

    _table_header(ws, 4, [
        "№", "Вид роботи", "Обсяг", "Од.", "Ставка USD", "Сума USD", "Сума грн"
    ])

    row = 5
    total_uah = 0.0
    total_usd = 0.0

    for i, work in enumerate(result["works"], 1):
        if work["price_usd"] == 0:
            continue  # пропускаємо нульові рядки
        is_alt = (i % 2 == 0)
        _data_row(ws, row, [
            i,
            work["name"],
            work["qty"],
            work["unit"],
            round(work["price_usd"], 2),
            round(work["total_usd"], 2),
            round(work["total_uah"], 2),
        ], is_alt)
        for col in [3, 5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        total_usd += work["total_usd"]
        total_uah += work["total_uah"]
        row += 1

    _total_row(ws, row, [
        "", "РАЗОМ РОБОТИ", "", "", "",
        round(total_usd, 2), round(total_uah, 2)
    ], CLR_TOTAL_ROW)
    for col in [6, 7]:
        ws.cell(row=row, column=col).number_format = '#,##0.00'

    _set_col_widths(ws, {"A": 5, "B": 42, "C": 10, "D": 10, "E": 13, "F": 13, "G": 14})
    ws.row_dimensions[4].height = 30

    return ws


# ════════════════════════════════════════════════════════════
# ЛИСТ 3 — ПІДСУМОК
# ════════════════════════════════════════════════════════════
def _sheet_summary(wb, result):
    ws = wb.create_sheet("Підсумок")
    rate = result["usd_rate"]
    now  = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Заголовок
    ws.merge_cells("A1:D1")
    ws["A1"] = "ЗАГАЛЬНИЙ КОШТОРИС"
    ws["A1"].font = Font(bold=True, size=15, color=CLR_HEADER_BLUE, name="Arial")
    ws["A1"].alignment = _align("center")

    ws.merge_cells("A2:D2")
    ws["A2"] = result["object_name"]
    ws["A2"].font = Font(bold=True, size=12, name="Arial")
    ws["A2"].alignment = _align("center")

    ws.merge_cells("A3:D3")
    ws["A3"] = f"Дата: {now} | Курс НБУ (Монобанк): 1 USD = {rate:.2f} грн"
    ws["A3"].font = _font(size=9, color="555555")
    ws["A3"].alignment = _align("center")

    # Інформація про об'єкт
    row = 5
    _section_header(ws, row, 1, 4, "ПАРАМЕТРИ ОБ'ЄКТУ")
    row += 1

    info_rows = [
        ("Система утеплення",   f"{result['system_name']} {result['insulation_thick']}мм"),
        ("Фарба",               result["paint_name"]),
        ("Площа фасаду (нетто)", f"{result['facade_area']} м²"),
        ("Площа фасаду (брутто)", f"{result['gross_area']} м²"),
        ("Площа відкосів",      f"{result['reveal_area']} м²"),
        ("Периметр будівлі",    f"{result['perimeter']} м"),
        ("Висота будівлі",      f"{result['building_height']} м"),
        ("Зовнішні кути",       f"{result['corner_count']} шт"),
        ("Вікна",               f"{result['window_count']} шт"),
        ("Двері",               f"{result['door_count']} шт"),
        ("Відливи",             f"{result['windowsill_lm']} п.м."),
        ("Стартовий профіль",   f"{result['start_profile_lm']} п.м."),
        ("Кутовий профіль",     f"{result['corner_lm_total']} п.м."),
        ("Ліси",                "так" if result["need_scaffold"] else "ні"),
    ]

    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = _header_fill(CLR_LIGHT_GRAY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = _font(size=10)
        ws.cell(row=row, column=2).alignment = _align("left")
        row += 1

    # Фінансовий підсумок
    row += 1
    _section_header(ws, row, 1, 4, "ФІНАНСОВИЙ ПІДСУМОК")
    row += 1

    _table_header(ws, row, ["Стаття витрат", "USD", "грн", ""])
    row += 1

    finance_rows = [
        ("Матеріали",   result["mat_total_usd"],   result["mat_total_uah"]),
        ("Роботи",      result["work_total_usd"],   result["work_total_uah"]),
        ("Доставка",    result["delivery_usd"],      result["delivery_uah"]),
    ]

    for i, (label, usd, uah) in enumerate(finance_rows):
        is_alt = (i % 2 == 0)
        fill = _header_fill(CLR_LIGHT_GRAY) if is_alt else _header_fill(CLR_WHITE)
        ws.cell(row=row, column=1, value=label).font = _font(size=10)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=round(usd, 2)).font = _font(size=10)
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = _align("right")
        ws.cell(row=row, column=3, value=round(uah, 2)).font = _font(size=10)
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).alignment = _align("right")
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = _border()
        row += 1

    # Фінальний підсумок
    _total_row(ws, row, [
        "ЗАГАЛЬНА СУМА",
        round(result["grand_total_usd"], 2),
        round(result["grand_total_uah"], 2),
        ""
    ], CLR_GRAND_TOTAL)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, name="Arial")
    for col in [2, 3]:
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=col).font = Font(bold=True, size=12, name="Arial")

    _set_col_widths(ws, {"A": 32, "B": 15, "C": 16, "D": 5})

    return ws


# ════════════════════════════════════════════════════════════
# ГОЛОВНА ФУНКЦІЯ — генерує Excel і повертає bytes
# ════════════════════════════════════════════════════════════
def generate_excel(result: dict) -> bytes:
    """
    Генерує Excel кошторис з 3 листами.
    Повертає bytes — готовий файл для відправки через Telegram.
    """
    wb = Workbook()

    # Видаляємо дефолтний лист
    default = wb.active
    wb.remove(default)

    _sheet_materials(wb, result)
    _sheet_works(wb, result)
    _sheet_summary(wb, result)

    # Встановлюємо активним лист Підсумок
    wb.active = wb["Підсумок"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def get_filename(result: dict) -> str:
    """Генерує ім'я файлу для кошторису"""
    date_str = datetime.now().strftime("%Y%m%d")
    # Беремо перші 30 символів назви об'єкту, замінюємо пробіли
    name = result["object_name"][:30].replace(" ", "_").replace("/", "-")
    return f"Кошторис_{name}_{date_str}.xlsx"
